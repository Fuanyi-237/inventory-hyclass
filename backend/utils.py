import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session
from .models import Item, Transaction, ItemState
from datetime import datetime
from fastapi import BackgroundTasks

# --- Excel Export ---
def export_inventory_to_excel(db: Session, file_path: str = "inventory_report.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = [
        "ID", "Unique ID", "Name", "Description", "Category", "State", "Location", "Purchase Date", "Expiry Date"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C5D9F1", fill_type="solid")

    items = db.query(Item).all()
    for item in items:
        ws.append([
            item.id, item.unique_id, item.name, item.description, item.category_id, item.state.value,
            item.location, item.purchase_date, item.expiry_date
        ])
    wb.save(file_path)
    return file_path

# --- Notification (stub) ---
def send_notification_email(to_email: str, subject: str, message: str):
    # Integrate with aiosmtplib or other mailer in production
    print(f"Email to {to_email}: {subject}\n{message}")
    return True

def schedule_notification(background_tasks: BackgroundTasks, to_email: str, subject: str, message: str):
    background_tasks.add_task(send_notification_email, to_email, subject, message)
